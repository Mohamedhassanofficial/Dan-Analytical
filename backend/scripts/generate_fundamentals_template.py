"""
Generate the blank Excel template admins fill in to bulk-upload stock
fundamentals (the 14 indicators per PPTX slide 83 + PPTX slide 106's
Yahoo Finance mapping).

Output:  backend/assets/templates/stock_fundamentals_template.xlsx

Layout:
  Row 1 — column headers (the only row the parser uses; column order is
          flexible, missing columns are skipped).
  Row 2 — human-readable hint row describing each column (parser ignores it
          because it starts with "#").
  Row 3 — a filled example row for 2222.SR (Saudi Aramco) with realistic
          numbers so the admin has a reference.
  Rows 4+ — admin fills these in; one per stock, uniqueness by `Symbol`.

Run:  python -m scripts.generate_fundamentals_template
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_PATH = (
    Path(__file__).resolve().parent.parent
    / "assets" / "templates" / "stock_fundamentals_template.xlsx"
)


# (column label, hint shown in row 2, Aramco example value in row 3)
COLUMNS: list[tuple[str, str, object]] = [
    ("Symbol",             "# 4-digit Tadawul code (e.g. 2222)",            "2222"),
    # Risk indicators
    ("Beta",               "# decimal, e.g. 0.85",                          0.85),
    ("Daily Volatility",   "# decimal, e.g. 0.0175 = 1.75%",                0.0175),
    ("Annual Volatility",  "# decimal, e.g. 0.1243 = 12.43%",               0.1243),
    ("Sharp Ratio",        "# decimal, e.g. -0.03",                         -0.03),
    ("VaR 1-Day",          "# positive decimal loss, e.g. 0.025 = 2.5%",    0.025),
    ("CAPM Expected Return", "# annual decimal, e.g. 0.08 = 8%",            0.08),
    # Financial indicators
    ("P/E",                "# e.g. 17.09",                                  17.09),
    ("EPS",                "# SR per share, e.g. 1.56",                     1.56),
    ("Dividend Yield",     "# decimal, e.g. 0.0517 = 5.17%",                0.0517),
    ("Annual Dividend Rate", "# SR per share, e.g. 1.33",                   1.33),
    ("ROE",                "# decimal, e.g. 0.2172 = 21.72%",               0.2172),
    ("Market to Book",     "# e.g. 4.18",                                   4.18),
    ("FCF Yield",          "# decimal, e.g. 0.0613 = 6.13%",                0.0613),
    ("Leverage",           "# decimal Debt/Equity, e.g. 0.21",              0.21),
    # Price snapshot
    ("Last Price",         "# SR, e.g. 40.15",                              40.15),
    ("Last Price Date",    "# ISO date, e.g. 2026-04-24",                   "2026-04-24"),
]


def build() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Fundamentals"

    header_fill = PatternFill("solid", fgColor="152F66")  # brand navy
    header_font = Font(color="FFFFFF", bold=True)
    hint_font = Font(color="4A6487", italic=True, size=9)
    example_font = Font(color="0A1A3D")

    for col_idx, (label, hint, example) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)

        # Row 1: header
        c = ws.cell(row=1, column=col_idx, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

        # Row 2: hint (prefixed with "#" so the parser skips it)
        h = ws.cell(row=2, column=col_idx, value=hint)
        h.font = hint_font
        h.alignment = Alignment(horizontal="left", vertical="center")

        # Row 3: Aramco example
        e = ws.cell(row=3, column=col_idx, value=example)
        e.font = example_font

        # Rough width — label length + 4 chars.
        ws.column_dimensions[letter].width = max(14, len(label) + 4)

    ws.freeze_panes = "B4"  # keep Symbol column + header+hint rows visible

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    return OUT_PATH


if __name__ == "__main__":
    p = build()
    print(f"✓ wrote {p} ({p.stat().st_size:,} bytes, {len(COLUMNS)} columns)")
